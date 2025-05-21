import os
from openpyxl import load_workbook, Workbook
from tkinter import Tk
from tkinter.filedialog import askdirectory

def generate_hyperlink_report():
    """
    Walks through a directory of AST tool outputs, checks hyperlinks in all Excel spreadsheets, and generates a report.
    The report includes the parent folder name, workbook name, whether a maps folder exists, whether hyperlinks are fixed,
    and a sample of hyperlinks.
    """
    # Prompt user to select the directory
    Tk().withdraw()  # Hide the root Tkinter window
    base_directory = askdirectory(title="Select the directory containing AST tool output folders")
    if not base_directory:
        print("No directory selected. Exiting.")
        return

    print(f"Starting hyperlink quality check in base directory: {base_directory}")

    # Prepare the report workbook
    report_wb = Workbook()
    report_ws = report_wb.active
    report_ws.title = "Hyperlink Report"
    report_ws.append(["Output Folder Name", "Workbook Name", "Maps Folder", "Hyperlinks Fixed (Yes/No)", "Sample Hyperlinks"])

    # Walk through all folders and subfolders
    for root, dirs, files in os.walk(base_directory):
        # Skip the base directory itself, process only subfolders
        if root == base_directory:
            continue

        # Get the parent folder name (next level down from the base directory)
        parent_folder_name = os.path.basename(root)

        # Check if the "maps" or "Maps" folder exists in the current directory
        maps_folder_exists = any(folder.lower() == "maps" for folder in dirs)

        print(f"Processing folder: {parent_folder_name}")

        # Process each Excel workbook in the current folder
        for file in files:
            if file.endswith(".xlsx"):
                workbook_path = os.path.join(root, file)
                workbook_name = os.path.basename(workbook_path)

                print(f"Checking workbook: {workbook_name}")

                try:
                    wb = load_workbook(workbook_path)
                    print(f"Workbook loaded successfully: {workbook_name}")
                except Exception as e:
                    print(f"Error loading workbook {workbook_name}: {e}")
                    continue

                # Check hyperlinks in the workbook
                hyperlinks_fixed = True
                has_hyperlinks = False  # Track if any hyperlinks are found
                sample_hyperlinks = []

                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.hyperlink:  # Check if the cell contains a hyperlink
                                has_hyperlinks = True  # At least one hyperlink exists
                                abs_path = cell.hyperlink.target
                                if not abs_path or not isinstance(abs_path, str):
                                    print(f"Warning: Invalid hyperlink found in {workbook_name}, sheet {sheet}. Skipping.")
                                    continue

                                # Normalize the hyperlink path and check if it is relative
                                normalized_path = abs_path.replace("\\", "/").strip()  # Normalize path to use forward slashes
                                if not normalized_path.startswith("maps/"):
                                    hyperlinks_fixed = False

                                # Collect a sample of the first 5 hyperlinks (normalized for consistency)
                                if len(sample_hyperlinks) < 5:
                                    sample_hyperlinks.append(normalized_path)

                # If no hyperlinks are found, mark as "No" and add a note
                if not has_hyperlinks:
                    hyperlinks_fixed = False
                    sample_hyperlinks.append("No hyperlinks found")

                # Add the results to the report
                report_ws.append([
                    parent_folder_name,
                    workbook_name,
                    "Yes" if maps_folder_exists else "No",
                    "Yes" if hyperlinks_fixed else "No",
                    ", ".join(sample_hyperlinks)
                ])

    # Save the report in the base directory
    report_name = os.path.basename(base_directory) + "_report.xlsx"
    report_path = os.path.join(base_directory, report_name)
    try:
        report_wb.save(report_path)
        print(f"Report saved successfully: {report_path}")
    except Exception as e:
        print(f"Error saving report: {e}")

    print("Hyperlink quality check completed.")
    
    # # Add this at the end of the script after saving the report
    # report_folder = os.path.dirname(report_path)  # Get the folder containing the report
    # try:
    #     print(f"Opening folder containing the report: {report_folder}")
    #     os.startfile(report_folder)  # Automatically open the folder in File Explorer (Windows)
    # except Exception as e:
    #     print(f"Error opening folder: {e}")

# Example usage
if __name__ == "__main__":
    generate_hyperlink_report()