import openpyxl


# Create the path
path = r"\\spatialfiles.bcgov\work\srm\nel\Local\Geomatics\Workarea\csostad\GitHubAutoAST\gss_authorizations\autoast\jobs_to_fail.xlsx"

# Load the workbook
wb = openpyxl.load_workbook(path)
print("Workbook loaded")

# Load the correct worksheet
ws = wb.active
print("Worksheet loaded")




header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
print(header)
    


# Check if 'AST CONDITION COLUMN' exists in the header
if "ast_condition" not in header:
    raise ValueError(f"Past_condition column not found in the spreadsheet.")
else:
    print("ast_condition column found")
    
    # Check if 'AST CONDITION COLUMN' exists in the header
if "dont_overwrite_outputs" not in header:
    raise ValueError(f"dont_overwite_outputs column not found in the spreadsheet.")
else:
    print("dont_overwite_outputs column found")



# Find the ast condition column and assign it to the correct index
ast_condition_index = header.index("ast_condition")  # +1 because Excel columns are 1-indexed
print(ast_condition_index)



# # Find the dont_overwrite_outputs column and assign it to the correct index
# dont_overwrite_outputs_index = header.index(self.DONT_OVERWRITE_OUTPUTS) + 1  # +1 because Excel columns are 1-indexed

# # Calculate the actual row index in Excel, +2 to account for header and 0-index
# excel_row_index = job_index + 1  # NOTE THIS COULD BE WHERE INDEX ISSUES ARE #BUG

# # Check if the row is blank before updating
# row_values = [ws.cell(row=excel_row_index, column=col).value for col in range(1, len(header) + 1)]
# if all(value is None or str(value).strip() == '' for value in row_values):
#     print(f"Row {excel_row_index} is blank, not updating.")
#     self.logger.info(f"Add_Job_Result -Job {job_index} / Row {excel_row_index} *** SHOULD THIS BE - 1? is blank, not updating.")
#     return  # Do not update if the row is blank

# # Update the condition for the specific job
# ws.cell(row=excel_row_index, column=ast_condition_index, value=condition)

# # if the condition in AST_CONDITION_COLUMN is 'Requeued" then go to the dont overwrite output column and change false to true
# if condition == 'Requeued':
#     print(f"Add Job Result - Job {job_index} failed, updating condition to 'Requeued'.")
#     self.logger.info(f"Add Job Result - Job {job_index} (Row {excel_row_index}) SHOULD THIS BE - 1? ***failed, updating condition to 'Requeued'.") 
#     ws.cell(row=excel_row_index, column=dont_overwrite_outputs_index, value="True")
#     self.logger.info(f"Add Job Result - Job {job_index} (Row {excel_row_index}) SHOULD THIS BE - 1?**** failed, updating dont_overwrite_outputs to 'True'.")

# # Save the workbook with the updated condition
# self.logger.info(f"Add Job Result - Updated Job {job_index} with condition '{condition}'.")
# wb.save(self.queuefile)
# self.logger.info(f"Add Job Result - Saving Workbook with updated condition")
# print(f"Updated row {excel_row_index} with condition '{condition}'.")