import os
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askdirectory

def change_hyperlinks_in_directory():
    """
    Updates hyperlinks in all Excel workbooks within a directory and its subdirectories to use relative paths.
    Prompts the user to select the directory containing multiple AST tool output folders.
    """
    # Prompt user to select the directory
    Tk().withdraw()  # Hide the root Tkinter window
    base_directory = askdirectory(title="Select the directory containing AST tool output folders")
    if not base_directory:
        print("No directory selected. Exiting.")
        return

    print(f"Starting hyperlink update process in base directory: {base_directory}")

    # Walk through all folders and subfolders
    for root, dirs, files in os.walk(base_directory):
        # Check if the "maps" folder exists in the current directory
        maps_folder = os.path.join(root, "maps")
        if not os.path.exists(maps_folder):
            continue  # Skip folders without a "maps" subfolder

        print(f"Processing folder: {root}")

        # Process each Excel workbook in the current folder
        for file_name in files:
            if file_name.endswith(".xlsx"):  # Check if the file is an Excel workbook
                workbook_path = os.path.join(root, file_name)
                print(f"Processing workbook: {file_name}")

                # Load the workbook
                try:
                    wb = load_workbook(workbook_path)
                    print(f"Workbook loaded successfully: {file_name}")
                except PermissionError:
                    print(f"Error: Workbook {file_name} is locked or in use. Please close it and try again.")
                    continue
                except Exception as e:
                    print(f"Error loading workbook {file_name}: {e}")
                    continue

                # Check if the workbook has sheets
                if not wb.sheetnames:
                    print(f"Warning: Workbook {file_name} has no sheets. Skipping.")
                    continue

                # Flag to track if all hyperlinks are already relative
                all_links_relative = True

                # Iterate through all sheets in the workbook
                for sheet in wb.sheetnames:
                    print(f"Processing sheet: {sheet}")
                    ws = wb[sheet]

                    # Iterate through all cells in the worksheet
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.hyperlink:  # Check if the cell contains a hyperlink
                                # Get the hyperlink target
                                abs_path = cell.hyperlink.target
                                if not abs_path or not isinstance(abs_path, str):
                                    print(f"Warning: Invalid hyperlink found in {file_name}, sheet {sheet}. Skipping.")
                                    continue

                                print(f"Found hyperlink: {abs_path}")

                                # Normalize the paths for comparison
                                normalized_maps_folder = os.path.normpath(maps_folder).lower()
                                normalized_abs_path = os.path.normpath(abs_path).lower()

                                # Check if the hyperlink points to the maps folder
                                if normalized_abs_path.startswith(normalized_maps_folder):
                                    # Convert the absolute path to a relative path
                                    rel_path = os.path.relpath(abs_path, start=root)
                                    print(f"Converting to relative path: {rel_path}")

                                    # Update the hyperlink to use the relative path
                                    cell.hyperlink = rel_path
                                    all_links_relative = False
                                else:
                                    print(f"Hyperlink does not point to maps folder: {abs_path}")

                # If all links were already relative, print a message and skip saving
                if all_links_relative:
                    print(f"The hyperlinks tool has already been run for {file_name}. All links are relative.")
                else:
                    # Save the updated workbook
                    try:
                        wb.save(workbook_path)
                        print(f"Updated hyperlinks saved in: {workbook_path}")
                    except PermissionError:
                        print(f"Error: Cannot save workbook {file_name}. It may be read-only.")
                    except Exception as e:
                        print(f"Error saving workbook {file_name}: {e}")

    print("Hyperlink update process completed.")

# Example usage
if __name__ == "__main__":
    change_hyperlinks_in_directory()