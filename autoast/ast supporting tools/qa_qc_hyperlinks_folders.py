import os
from openpyxl import load_workbook, Workbook
from tkinter import Tk
from tkinter.filedialog import askdirectory

def generate_hyperlink_report():
    """
    Walks through a directory of AST tool outputs, checks hyperlinks in all Excel spreadsheets, and generates a report.
    The report includes the parent folder name, workbook name, sheet name, whether a maps folder exists, whether hyperlinks are fixed,
    a sample of hyperlinks, and the "Status Run By / Date" field from automated_status_sheet.xlsx.
    """
    # Prompt user to select the directory
    Tk().withdraw()  # Hide the root Tkinter window
    base_directory = askdirectory(title="Select the directory containing AST tool output folders")
    if not base_directory:
        print("No directory selected. Exiting.")
        return

    print(f"Starting hyperlink quality check in base directory: {base_directory}")

    # Prepare the report workbook
    report_wb = Workbook()
    report_ws = report_wb.active
    report_ws.title = "Hyperlink Report"
    
    # Initialize the report header
    report_ws.append([
        "Output Folder Name",
        "Workbook Name",
        "Sheet Name",  # Add the sheet name column
        "Status Run By / Date",  # Add the new column
        "Maps Folder",
        "Hyperlinks Fixed (Yes/No)",
        "Sample Hyperlinks"
    ])

    # Walk through all folders and subfolders
    for root, dirs, files in os.walk(base_directory):
        # Skip the base directory itself, process only subfolders
        if root == base_directory:
            continue

        # Get the parent folder name (next level down from the base directory)
        parent_folder_name = os.path.basename(root)

        # Check if the "maps" or "Maps" folder exists in the current directory
        maps_folder_exists = any(folder.lower() == "maps" for folder in dirs)

        print(f"Processing folder: {parent_folder_name}")

        # Process each Excel workbook in the current folder
        for file in files:
            if file.endswith(".xlsx"):
                workbook_path = os.path.join(root, file)
                workbook_name = os.path.basename(workbook_path)

                print(f"Checking workbook: {workbook_name}")

                try:
                    wb = load_workbook(workbook_path)
                    print(f"Workbook loaded successfully: {workbook_name}")
                except Exception as e:
                    print(f"Error loading workbook {workbook_name}: {e}")
                    continue

                # Extract "Status Run By / Date" from automated_status_sheet.xlsx
                status_run_by_date = "N/A"  # Default value if not found
                if workbook_name.lower() == "automated_status_sheet.xlsx":  # Case-insensitive check for the workbook name
                    try:
                        # Access the first sheet (Crown Land Status)
                        sheet1 = wb["Crown Land Status"]

                        # Get the value of cell B10
                        cell_value = sheet1["B10"].value

                        # Print the value
                        print(f"The value of cell B10 is: {cell_value}")

                        # Assign the value to the "Status Run By / Date" field
                        status_run_by_date = cell_value if cell_value else "N/A"

                    except Exception as e:
                        print(f"Error reading 'Status Run By / Date' from {workbook_name}: {e}")
        
                # Check hyperlinks in the workbook
                hyperlinks_fixed = True
                has_hyperlinks = False  # Track if any hyperlinks are found
                sample_hyperlinks = []

                for sheet in wb.sheetnames:
                    ws = wb[sheet]
                    for row in ws.iter_rows():
                        for cell in row:
                            if cell.hyperlink:  # Check if the cell contains a hyperlink
                                has_hyperlinks = True  # At least one hyperlink exists
                                abs_path = cell.hyperlink.target
                                if not abs_path or not isinstance(abs_path, str):
                                    print(f"Warning: Invalid hyperlink found in {workbook_name}, sheet {sheet}. Skipping.")
                                    continue

                                # Normalize the hyperlink path
                                normalized_path = abs_path.replace("\\", "/").strip()

                                # Check if the hyperlink is relative
                                if normalized_path.startswith("maps/"):
                                    hyperlinks_fixed = True  # Mark as fixed if all hyperlinks are relative
                                else:
                                    hyperlinks_fixed = False  # Mark as not fixed if any absolute hyperlink is found

                                # Collect all hyperlinks (normalized for consistency)
                                sample_hyperlinks.append(normalized_path)

                    # Add the results to the report
                    report_ws.append([
                        parent_folder_name,
                        workbook_name,
                        sheet,  # Add the sheet name to the report
                        status_run_by_date,  # Add the "Status Run By / Date" value
                        "Yes" if maps_folder_exists else "No",
                        "Yes" if hyperlinks_fixed else "No",
                        ", ".join(sample_hyperlinks)  # Record all hyperlinks in the sample column
                    ])

    # Save the report in the base directory
    report_name = os.path.basename(base_directory) + "_report.xlsx"
    report_path = os.path.join(base_directory, report_name)

    try:
        report_wb.save(report_path)
        print(f"Report saved successfully: {report_path}")
    except Exception as e:
        print(f"Error saving report: {e}")

    print("Hyperlink quality check completed.")

    # Open the report file automatically
    try:
        # Normalize the path to ensure compatibility with Windows
        normalized_report_path = os.path.normpath(report_path)
        print(f"Opening report: {normalized_report_path}")
        os.startfile(normalized_report_path)  # Automatically open the report file in Excel
    except Exception as e:
        print(f"Error opening report: {e}")

# Example usage
if __name__ == "__main__":
    generate_hyperlink_report()