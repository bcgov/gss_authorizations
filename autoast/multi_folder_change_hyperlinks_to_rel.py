import os
from openpyxl import load_workbook
from tkinter import Tk
from tkinter.filedialog import askdirectory

def change_hyperlinks_in_directory():
    """
    Updates hyperlinks in all specified Excel workbooks within a directory and its subdirectories to use relative paths.
    Prompts the user to select the directory containing multiple AST tool output folders.
    """
    # Prompt user to select the directory
    Tk().withdraw()  # Hide the root Tkinter window
    base_directory = askdirectory(title="Select the directory containing AST tool output folders")
    if not base_directory:
        print("No directory selected. Exiting.")
        return

    print(f"Starting hyperlink update process in base directory: {base_directory}")

    # Define the names of the workbooks to process
    workbooks_to_process = [
        "automated_status_sheet.xlsx",
        "one_status_common_dataset_aoi.xlsx",
        "one_status_tabs_1_and_2.xlsx"
    ]

    # Walk through all folders and subfolders
    for root, dirs, files in os.walk(base_directory):
        # Check if the "maps" folder exists in the current directory
        maps_folder = os.path.join(root, "maps")
        if not os.path.exists(maps_folder):
            continue  # Skip folders without a "maps" subfolder

        print(f"Processing folder: {root}")

        # Process each workbook in the current folder
        for workbook_name in workbooks_to_process:
            workbook_path = os.path.join(root, workbook_name)
            if not os.path.exists(workbook_path):
                continue  # Skip if the workbook does not exist

            print(f"Processing workbook: {workbook_name}")

            # Load the workbook
            try:
                wb = load_workbook(workbook_path)
                print(f"Workbook loaded successfully: {workbook_name}")
            except PermissionError:
                print(f"Error: Workbook {workbook_name} is locked or in use. Please close it and try again.")
                continue
            except Exception as e:
                print(f"Error loading workbook {workbook_name}: {e}")
                continue

            # Check if the workbook has sheets
            if not wb.sheetnames:
                print(f"Warning: Workbook {workbook_name} has no sheets. Skipping.")
                continue

            # Flag to track if all hyperlinks are already relative
            all_links_relative = True

            # Iterate through all sheets in the workbook
            for sheet in wb.sheetnames:
                print(f"Processing sheet: {sheet}")
                ws = wb[sheet]

                # Iterate through all cells in the worksheet
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.hyperlink:  # Check if the cell contains a hyperlink
                            # Get the hyperlink target
                            abs_path = cell.hyperlink.target
                            if not abs_path or not isinstance(abs_path, str):
                                print(f"Warning: Invalid hyperlink found in {workbook_name}, sheet {sheet}. Skipping.")
                                continue

                            print(f"Found hyperlink: {abs_path}")

                            # Check if the hyperlink is already relative
                            if abs_path.startswith("maps/") or abs_path.startswith("maps\\"):
                                print(f"Hyperlink is already a relative path: {abs_path}")
                                continue  # Skip processing this hyperlink

                            # Ensure the hyperlink points to the maps folder
                            if maps_folder.lower() in abs_path.lower():  # Case-insensitive check
                                # Convert the absolute path to a relative path
                                rel_path = os.path.relpath(abs_path, start=root)
                                print(f"Converting to relative path: {rel_path}")

                                # Update the hyperlink to use the relative path
                                cell.hyperlink = rel_path
                                all_links_relative = False
                            else:
                                print(f"Hyperlink does not point to maps folder: {abs_path}")

            # If all links were already relative, print a message and skip saving
            if all_links_relative:
                print(f"The hyperlinks tool has already been run for {workbook_name}. All links are relative.")
            else:
                # Save the updated workbook
                try:
                    wb.save(workbook_path)
                    print(f"Updated hyperlinks saved in: {workbook_path}")
                except PermissionError:
                    print(f"Error: Cannot save workbook {workbook_name}. It may be read-only.")
                except Exception as e:
                    print(f"Error saving workbook {workbook_name}: {e}")

    print("Hyperlink update process completed.")

# Example usage
if __name__ == "__main__":
    change_hyperlinks_in_directory()