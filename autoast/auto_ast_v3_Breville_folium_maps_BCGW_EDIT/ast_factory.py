import os
from openpyxl import Workbook, load_workbook
import arcpy
import logging
import traceback
import multiprocessing as mp
from mp_worker import process_job_mp
from aoi_utilities import build_aoi_from_shp
from aoi_utilities import build_aoi_from_kml

class AST_FACTORY:
    ''' AST_FACTORY creates and manages status tool runs '''
    XLSX_SHEET_NAME = 'ast_config'
    AST_PARAMETERS = {
        0: 'region',
        1: 'feature_layer',
        2: 'crown_file_number',
        3: 'disposition_number',
        4: 'parcel_number',
        5: 'output_directory',
        6: 'output_directory_same_as_input',
        7: 'dont_overwrite_outputs',
        8: 'skip_conflicts_and_constraints',
        9: 'suppress_map_creation',
        10: 'add_maps_to_current',
        11: 'run_as_fcbc',
    }
    
    ADDITIONAL_PARAMETERS = {
        12: 'ast_condition',
        13: 'file_number'
    }
    
    AST_CONDITION_COLUMN = 'ast_condition'
    DONT_OVERWRITE_OUTPUTS = 'dont_overwrite_outputs'
    AST_SCRIPT = ''
    job_index = None  # Initialize job_index as a global variable
    
    def __init__(self, queuefile, logger, current_path, connection_path):
        self.queuefile = queuefile
        self.logger = logger
        self.current_path = current_path
        self.connection_path = connection_path

    def load_jobs(self):
        '''
        load jobs will check for the existence of the queuefile,
        if it exists it will load the jobs from the queuefile.
        It checks if they are Complete and if not, adds them as Queued.
        '''
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Loading Jobs...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")

        self.jobs = []

        assert os.path.exists(self.queuefile), "Queue file does not exist"
        if os.path.exists(self.queuefile):
            try:
                wb = load_workbook(filename=self.queuefile)
                ws = wb[self.XLSX_SHEET_NAME]

                header = list(next(ws.iter_rows(min_row=1, max_row=1, values_only=True)))
                
                data = []
                for row in ws.iter_rows(min_row=2, max_col=None, values_only=True):
                    print(f'Row is {row}')
                    data.append(row)
                
                for job_index, row_data in enumerate(data):  
                    self.logger.info("\n")
                    self.logger.info("-------------------------------------------------------------------------------")
                    self.logger.info(f"-                        Load Jobs: Start of Job {job_index}                               -")
                    self.logger.info("-------------------------------------------------------------------------------")
                    self.logger.info("\n")
                    
                    job = {}
                    self.logger.info('Load Jobs - Creating empty dictionary')
                    ast_condition = None  
                        
                    if all((value is None or str(value).strip() == '') for value in row_data):
                        print(f"Load Jobs - Skipping blank row at job index ({job_index}) ")
                        self.logger.info(f"Load Jobs - Skipping blank row at index ({job_index}) ")
                        continue

                    for key, value in zip(header, row_data):
                        if key is not None and key.lower() == self.AST_CONDITION_COLUMN.lower():
                            ast_condition = value if value is not None else ""
                        value = "" if value is None else value
                        if key is not None:
                            job[key] = value

                    if ast_condition and ast_condition.upper() == 'COMPLETE':
                        print(f"Skipping job {job_index} as it is marked COMPLETE.")
                        self.logger.info(f"Load Jobs - Skipping job {job_index} as it is marked COMPLETE.")
                    else:
                        if ast_condition is None or ast_condition.strip() == '' or ast_condition.upper() != 'COMPLETE':
                            ast_condition = 'Queued'
                            job[self.AST_CONDITION_COLUMN] = ast_condition
                            self.logger.info(f"Load Jobs - (Queued assigned to Job ({job_index}) is ({ast_condition})")
                            try:
                                self.add_job_result(job_index, ast_condition)
                                self.logger.info(f"Load Jobs - Added job condition '{ast_condition}' for job {job_index} to jobs list")
                            except Exception as e:
                                print(f"Error updating Excel sheet at row {job_index}: {e}")
                                self.logger.error(f"Load Jobs - Error updating Excel sheet at row {job_index}: {e}")
                                continue

                            try:
                                self.logger.info(f"Classifying input type for job {job_index}")
                                self.classify_input_type(job)
                            except Exception as e:
                                print(f"Error classifying input type for job {job}: {e}")
                                self.logger.error(f"Error classifying input type for job {job}: {e}")
                    self.jobs.append(job)
                    print(f"Load Jobs - Job Condition for job ({job_index}) is not Complete: Writing ({ast_condition}) to ast_condition. Adding job: {job_index} to jobs list")
                    self.logger.info(f"Load Jobs - Job Condition is not Complete ({ast_condition}), adding job: {job_index} to jobs list")
                    self.logger.info("\n")
                    self.logger.info("-------------------------------------------------------------------------------")
                    self.logger.info(f"-                        End of Job {job_index}                                -")
                    self.logger.info("-------------------------------------------------------------------------------")
                    self.logger.info("\n")
                    
            except Exception as e:
                print(f"Unexpected error loading jobs: {e}")
                self.logger.error(f"Unexpected error loading jobs: {e}")

            return self.jobs

    def classify_input_type(self, job):
        '''Classify the input type and process accordingly.'''
        if job.get('feature_layer'):
            print(f'Feature layer found: {job["feature_layer"]}')
            self.logger.info(f'Classifying Input Type - Feature layer found: {job["feature_layer"]}')
            feature_layer_path = job['feature_layer']
            print(f"Processing feature layer: {feature_layer_path}")
            self.logger.info(f"Classifying Input Type - Processing feature layer: {feature_layer_path}")

            if feature_layer_path.lower().endswith('.kml'):
                print('KML found, building AOI from KML')
                self.logger.info('Classifying Input Type - KML found, building AOI from KML')
                # Pass logger to the function if needed
                job['feature_layer'] = build_aoi_from_kml(feature_layer_path, self.logger)
            elif feature_layer_path.lower().endswith('.shp'):
                if job.get('file_number'):
                    print(f"File number found, running FW setup on shapefile: {feature_layer_path}")
                    self.logger.info(f"Classifying Input Type - File number found, running FW setup on shapefile: {feature_layer_path}")
                    # If your build_aoi_from_shp function requires template and logger, adjust accordingly.
                    new_feature_layer_path = build_aoi_from_shp(job, feature_layer_path, None, self.logger)
                    job['feature_layer'] = new_feature_layer_path
                else:
                    print('No FW File Number provided for the shapefile, using original shapefile path')
                    self.logger.info('Classifying Input Type - No FW File Number provided, using original shapefile path')
            else:
                print(f"Unsupported feature layer format: {feature_layer_path}")
                self.logger.warning(f"Classifying Input Type - Unsupported feature layer format: {feature_layer_path} - Marking job as Failed")
                self.add_job_result(job, 'Failed')
        else:
            print('No feature layer provided in job')
            self.logger.warning('Classifying Input Type - No feature layer provided in job')

    def add_job_result(self, job_index, condition):
        ''' 
        Function adds result information to the Excel spreadsheet.
        '''
        print("Running Add Job Results...")
        self.logger.info("\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Running Add Job Results from Load Jobs Function")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info("\n")
     
        try:
            wb = load_workbook(filename=self.queuefile)
            self.logger.info("Add Job Result - Workbook loaded")
            ws = wb[self.XLSX_SHEET_NAME]
            header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            
            if self.AST_CONDITION_COLUMN not in header:
                raise ValueError(f"'{self.AST_CONDITION_COLUMN}' column not found in the spreadsheet.")
            
            self.logger.info(f"Add Job Result - '{self.AST_CONDITION_COLUMN}' column found in the spreadsheet.")
            ast_condition_index = header.index(self.AST_CONDITION_COLUMN) + 1
            dont_overwrite_outputs_index = header.index(self.DONT_OVERWRITE_OUTPUTS) + 1 if self.DONT_OVERWRITE_OUTPUTS in header else None
            excel_row_index = job_index + 2
            self.logger.info(f"Add Job Result - Calculated Excel row index as {excel_row_index} for job index {job_index}")
            
            row_values = []
            for col in range(1, len(header) + 1):
                cell_value = ws.cell(row=excel_row_index, column=col).value
                row_values.append(cell_value)
            if all(value is None or str(value).strip() == '' for value in row_values):
                print(f"Row {excel_row_index} is blank, not updating.")
                self.logger.info(f"Add_Job_Result -Job {job_index} / Row {excel_row_index} is blank")
                return

            ws.cell(row=excel_row_index, column=ast_condition_index, value=condition)

            if condition == 'Requeued' and dont_overwrite_outputs_index is not None:
                self.logger.info(f"Add Job Result - Job {job_index} (Row {excel_row_index}) updating condition to 'Requeued'.")
                ws.cell(row=excel_row_index, column=dont_overwrite_outputs_index, value="True")
                self.logger.info(f"Add Job Result - Job {job_index} (Row {excel_row_index}) updating dont_overwrite_outputs to 'True'.")
            
            self.logger.info(f"Add Job Result - Updated Job {job_index} with condition '{condition}'.")
            wb.save(self.queuefile)
            self.logger.info("Add Job Result - Saving Workbook with updated condition")
            print(f"Updated row {excel_row_index} with condition '{condition}'.")

        except Exception as e:
            print(f"Unexpected error while adding job result: {e}")
            self.logger.error(f"Unexpected error while adding job result: {e}")

    def batch_ast(self):
        '''
        Uses multiprocessing to run jobs in parallel.
        '''
        self.logger.info("\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Batch AST: Batching Jobs with Multiprocessing...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info("\n")
        
        JOB_TIMEOUT = 21600  # 6 hours in seconds
        self.logger.info(f"Batch Ast: Job Timeout set to {JOB_TIMEOUT} seconds")
        print(f"Batch Ast: Job Timeout set to {JOB_TIMEOUT} seconds")

        processes = []
        manager = mp.Manager()
        return_dict = manager.dict()

        for job_index, job in enumerate(self.jobs):
            self.logger.info(f"Batch Ast: Starting job {job_index}")
            print(f"Batch Ast: Starting job {job_index} Job ({job})")

            if job.get(self.AST_CONDITION_COLUMN) in ['Queued', 'Requeued']:
                p = mp.Process(target=process_job_mp, args=(self, job, job_index, self.current_path, self.connection_path, return_dict))
                processes.append((p, job_index))
                p.start()
                self.logger.info(f"Batch Ast: {job.get(self.AST_CONDITION_COLUMN)} Job {job_index}.....Multiprocessing started......")
                print("Batch Ast: Queued Job...Multiprocessing started......")

        timeout_failed_counter = 0
        success_counter = 0
        worker_failed_counter = 0
        other_exception_failed_counter = 0
        for process, job_index in processes:
            process.join(JOB_TIMEOUT)
            if process.is_alive():
                print(f"Batch Ast: Job {job_index} exceeded timeout. Terminating process.")
                self.logger.warning(f"Batch Ast: Job {job_index} exceeded timeout. Terminating process.")
                process.terminate()
                process.join()
                self.add_job_result(job_index, 'Failed') 
                timeout_failed_counter += 1
                self.logger.error(f"Batch Ast: Job {job_index} exceeded timeout. Marking as Failed. Failed counter is {timeout_failed_counter}")
            else:
                result = return_dict.get(job_index)
                if result == 'Success':
                    success_counter += 1
                    self.add_job_result(job_index, 'COMPLETE')
                    print(f"Batch Ast: Job {job_index} completed successfully.")
                    self.logger.info(f"Batch Ast: Job {job_index} completed successfully. Success counter is {success_counter}")
                elif result == 'Failed':
                    self.add_job_result(job_index, 'Failed')
                    worker_failed_counter += 1
                    print(f"Batch Ast: Job {job_index} failed due to an exception.")
                    self.logger.error(f"Batch AST: Job {job_index} failed due to an exception in the Worker. Other exception failed counter is {worker_failed_counter}")
                else:
                    self.add_job_result(job_index, 'Unknown Error')
                    other_exception_failed_counter += 1
                    print(f"Batch Ast: Job {job_index} failed with unknown status.")
                    self.logger.error(f"Batch AST: Job {job_index} failed with unknown status. Other Exception failed counter is {other_exception_failed_counter}")
         
        self.logger.info("\n")    
        self.logger.info("Batch Ast Complete - Check separate worker log file for more details")

    def re_load_failed_jobs_V2(self):
        '''
        Reload failed jobs will check for the existence of the queuefile and update job statuses.
        '''
        self.logger.info("\n")
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Re loading Failed Jobs V2.....")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")
        self.logger.info("\n")

        self.jobs = []

        assert os.path.exists(self.queuefile), "Queue file does not exist"
        if os.path.exists(self.queuefile):
            try:
                wb = load_workbook(filename=self.queuefile)
                ws = wb[self.XLSX_SHEET_NAME]
                
                print(f'Workbook loaded is {wb}')   
                self.logger.info(f'Re load Failed Jobs: Workbook loaded is {wb}') 
                
                header = list(next(ws.iter_rows(min_row=1, max_col=None, values_only=True)))
                data = []
                self.logger.info("Re load Failed Jobs: Reading all data rows and building data list")
                for row in ws.iter_rows(min_row=2, max_col=None, values_only=True):
                    print(f'Row is {row}')
                    data.append(row)

                self.logger.info("Re load Failed Jobs: Iterating over each row of data")
                for job_index, row_data in enumerate(data):
                    self.logger.info("\n")
                    self.logger.info("------------------------------------------------------------------------------------")
                    self.logger.info(f"-                        Re Load Failed Jobs: Start of Job {job_index}                               -")
                    self.logger.info("------------------------------------------------------------------------------------")
                    self.logger.info("\n")
                    
                    job = {}
                    self.logger.info('Re load Jobs - Creating empty dictionary')
                    ast_condition = ''
                        
                    if all((value is None or str(value).strip() == '') for value in row_data):
                        print(f"Re Load Failed Jobs: Skipping blank row at index {job_index}")
                        self.logger.info(f"Re Load Failed Jobs: Skipping blank row at index {job_index}")
                        continue

                    for key, value in zip(header, row_data):
                        if key is not None and key.lower() == self.AST_CONDITION_COLUMN.lower():
                            ast_condition = value if value is not None else ""
                        value = "" if value is None else value
                        if key is not None:
                            print("Re Load Failed Jobs: Assigning values to job dictionary")
                            job[key] = value

                    if ast_condition.upper() == 'COMPLETE':
                        print(f"Re Load Failed Jobs: Skipping job {job_index} as it is marked {ast_condition}.")
                        self.logger.info(f"Re Load Failed Jobs: Skipping job {job_index} as it is marked COMPLETE.")
                        ast_condition = 'COMPLETE'
                    elif ast_condition.upper() == 'FAILED':
                        self.logger.info(f"Re Load Failed Jobs: Requeuing {job_index} as it is marked Failed.")
                        ast_condition = 'Requeued'
                    else:
                        self.logger.warning(f"Re Load Failed Jobs: Job {job_index} is not marked as Complete or Failed. Please check the workbook. Skipping this job.")
                        ast_condition = 'ERROR'
                    
                    job[self.AST_CONDITION_COLUMN] = ast_condition
                        
                    print(f"Re Load Failed Jobs: Job {job_index} is marked as Failed, re-assigning ast condition to {ast_condition}")
                    self.logger.info(f"Re Load Failed Jobs: Job {job_index}'s ast condition has been updated as '{ast_condition}'")

                    try:
                        self.add_job_result(job_index, ast_condition)
                        self.logger.info(f"Re load Jobs - Added job condition '{ast_condition}' for job {job_index} to jobs list")
                        self.logger.info(f"Reload Failed Jobs - Saving Job Index ({job_index}) with new condition.")
                        wb.save(self.queuefile)
                    except Exception as e:
                        print(f"Error updating Excel sheet at row {job_index}: {e}")
                        self.logger.error(f"Re load Jobs - Error updating Excel sheet at row {job_index}: {e}")
                        self.logger.error(traceback.format_exc())
                        continue
                        
                    self.jobs.append(job)
                    print(f"Re load Jobs - Job Condition is not Complete ({ast_condition}), adding job: {job_index} to jobs list")
                    self.logger.info(f"Re load Jobs - Job Condition is not Complete ({ast_condition}), adding job: {job_index} to jobs list")
                    self.logger.info("\n")
                    self.logger.info("------------------------------------------------------------------------------------")   
                    self.logger.info(f" Job list is {job}")
                    self.logger.info("------------------------------------------------------------------------------------")
                    self.logger.info("\n")
                    
                    print(f"Re Load Jobs - Job dictionary is {job}")
                    self.logger.info(f"Re load Jobs - Job {job_index} dictionary is {job}")
                            
            except Exception as e:
                print(f"Unexpected error re loading jobs: {e}")
                self.logger.error(f"Re Load Failed Jobs Unexpected error loading jobs: {e}")
                self.logger.error(traceback.format_exc())

        return self.jobs   

    def create_new_queuefile(self):
        '''Write a new queuefile with preset header'''
        self.logger.info("##########################################################################################################################")
        self.logger.info("#")
        self.logger.info("Creating New Queuefile...")
        self.logger.info("#")
        self.logger.info("##########################################################################################################################")

        wb = Workbook()
        ws = wb.active
        ws.title = self.XLSX_SHEET_NAME
        headers = list(self.AST_PARAMETERS.values())
        headers.append(self.AST_CONDITION_COLUMN)
        for h in headers:
            c = headers.index(h) + 1
            ws.cell(row=1, column=c).value = h
        wb.save(self.queuefile)

    def capture_arcpy_messages(self):
        '''Reassigns the arcpy messages to the logger'''
        arcpy_messages = arcpy.GetMessages(0)
        arcpy_warnings = arcpy.GetMessages(1)
        arcpy_errors = arcpy.GetMessages(2)
        
        if arcpy_messages:
            self.logger.info(f'ast_toolbox arcpy messages: {arcpy_messages}')
        if arcpy_warnings:
            self.logger.warning(f'ast_toolbox arcpy warnings: {arcpy_warnings}')
        if arcpy_errors:
            self.logger.error(f'ast_toolbox arcpy errors: {arcpy_errors}')
