import os
from openpyxl import load_workbook

def change_hyperlinks(output_folder):
    """
    The change_hyperlinks script updates hyperlinks in all Excel workbooks within a given output folder to use relative paths instead of absolute paths.
    It validates the existence of the output folder and the `maps` subfolder before processing.
    For each workbook, it iterates through all sheets and cells, identifying hyperlinks that point to files in the `maps` folder.
    If a hyperlink is absolute, it converts it to a relative path (e.g., `maps/filename.pdf`).
    The script skips hyperlinks that are already relative and handles errors such as missing files, locked workbooks, or invalid hyperlinks,
    providing clear messages for each scenario.

    Arguments:
        output_folder (str): The path to the output folder containing the Excel workbooks and the maps folder.
    """
    print(f"Starting hyperlink update process in folder: {output_folder}")
    
    # Validate the output folder
    if not os.path.exists(output_folder):
        print(f"Error: Output folder does not exist: {output_folder}")
        return

    # Path to the maps folder
    maps_folder = os.path.join(output_folder, "maps")
    print(f"Maps folder path: {maps_folder}")

    # Check if the maps folder exists
    if not os.path.exists(maps_folder):
        print(f"Error: Maps folder does not exist at {maps_folder}")
        return

    # Iterate through all Excel workbooks in the folder
    for file_name in os.listdir(output_folder):
        if file_name.endswith(".xlsx"):  # Check if the file is an Excel workbook
            workbook_path = os.path.join(output_folder, file_name)
            print(f"Processing workbook: {file_name}")

            # Load the workbook
            try:
                wb = load_workbook(workbook_path)
                print(f"Workbook loaded successfully: {file_name}")
            except PermissionError:
                print(f"Error: Workbook {file_name} is locked or in use. Please close it and try again.")
                continue
            except Exception as e:
                print(f"Error loading workbook {file_name}: {e}")
                continue

            # Check if the workbook has sheets
            if not wb.sheetnames:
                print(f"Warning: Workbook {file_name} has no sheets. Skipping.")
                continue

            # Flag to track if all hyperlinks are already relative
            all_links_relative = True

            # Iterate through all sheets in the workbook
            for sheet in wb.sheetnames:
                print(f"Processing sheet: {sheet}")
                ws = wb[sheet]

                # Iterate through all cells in the worksheet
                for row in ws.iter_rows():
                    for cell in row:
                        if cell.hyperlink:  # Check if the cell contains a hyperlink
                            # Get the hyperlink target
                            abs_path = cell.hyperlink.target
                            if not abs_path or not isinstance(abs_path, str):
                                print(f"Warning: Invalid hyperlink found in {file_name}, sheet {sheet}. Skipping.")
                                continue

                            print(f"Found hyperlink: {abs_path}")

                            # Check if the hyperlink is already relative
                            if abs_path.startswith("maps/") or abs_path.startswith("maps\\"):
                                print(f"Hyperlink is already a relative path: {abs_path}")
                                continue  # Skip processing this hyperlink

                            # Ensure the hyperlink points to the maps folder
                            if maps_folder.lower() in abs_path.lower():  # Case-insensitive check
                                # Convert the absolute path to a relative path
                                rel_path = os.path.relpath(abs_path, start=output_folder)
                                print(f"Converting to relative path: {rel_path}")

                                # Update the hyperlink to use the relative path
                                cell.hyperlink = rel_path
                                all_links_relative = False
                            else:
                                print(f"Hyperlink does not point to maps folder: {abs_path}")

            # If all links were already relative, print a message and skip saving
            if all_links_relative:
                print(f"The hyperlinks tool has already been run for {file_name}. All links are relative.")
            else:
                # Save the updated workbook
                try:
                    wb.save(workbook_path)
                    print(f"Updated hyperlinks saved in: {workbook_path}")
                except PermissionError:
                    print(f"Error: Cannot save workbook {file_name}. It may be read-only.")
                except Exception as e:
                    print(f"Error saving workbook {file_name}: {e}")

    print("Hyperlink update process completed.")

# Example usage
if __name__ == "__main__":
    output_folder = r"\\spatialfiles.bcgov\work\srm\nel\Local\Geomatics\Workarea\SharedWork\gr_2025_26\136_south coast"  # Replace with the actual path to the output folder
    change_hyperlinks(output_folder)